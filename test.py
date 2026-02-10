import pandas as pd
from openpyxl import load_workbook

# Example race + entry data
df = pd.DataFrame({
    "race_name": ["Monaco GP"] * 2,
    "year": [2025] * 2,
    "race_date": ["25/05/2025"] * 2,
    "circuit_name": ["Circuit de Monaco"] * 2,
    "team_name": ["Ferrari", "Red Bull"],
    "driver_name": ["Leclerc", "Verstappen"],
    "current_points": [18, 25],
    "previous_year_winner": [False, True],
    "fastest_lap_driver": [False, True]
})

# Export to Excel
excel_file = "race_results.xlsx"
df.to_excel(excel_file, index=False)

# Load workbook and merge race-level columns
wb = load_workbook(excel_file)
ws = wb.active

# Merge cells for race-level columns across rows 2 and 3
ws.merge_cells("A2:A3")  # race_name
ws.merge_cells("B2:B3")  # year
ws.merge_cells("C2:C3")  # race_date
ws.merge_cells("D2:D3")  # circuit_name

wb.save(excel_file)
